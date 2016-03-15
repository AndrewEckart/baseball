import openpyxl

def export_data(players,destpath='brdata.xlsx',end=2015,numSeasons=1):
	print("Exporting to file: ",destpath)
	wb = openpyxl.Workbook()
	currentyear = end+1
	prev_seasons = []
# Create worksheets for current season
	hitters,pitchers = wb.active, wb.create_sheet()
	hitters.title,pitchers.title = ("%s Hitters" %currentyear,"%s Pitchers" %currentyear)
	hitters.append(['Player','Position','Team','Age'])
	pitchers.append(['Player','Position','Team','Age'])
# Create index of past seasons	
	for i in range(numSeasons): prev_seasons.append(str(end-i))
	for year in prev_seasons:
		h = wb.create_sheet()
		h.title = (year + "h")
		h.append(['Player','Team','Age','PA','AB','H','BB','HR','R','RBI','SB','BA','OBP'])
		p = wb.create_sheet()
		p.title = (year + "p")
		p.append(['Player','Team','Age','G','GS','W','L','IP','SV','H','BB','SO','ERA','WHIP'])
	print(wb.worksheets)	
	for player in players:
		isPitcher = ("Pitcher" in player.position)
		if isPitcher: pitchers.append([player.name,player.position,player.team])
		else: hitters.append([player.name,player.position,player.team])
		for year in prev_seasons:
			if hasattr(player, 'seasons') and year in player.seasons.keys():
				if isPitcher: ws = wb[(year + "p")]
				else: ws = wb[(year + "h")]
				try: ws.append(player.seasons[year].exportSeason())
				except: print('Error during export! Player:%s, Season: %s' %(player.name,year))
	wb.save(destpath)
	return